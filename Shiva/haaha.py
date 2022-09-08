'''
import docx
import random
doc = docx.Document("Descriptive_Python_GV.docx")
for i in doc.tables:
    row=1
    for j in i.rows:
        col = 1

        if len(j.cells) != 4:
            continue
        try:
            j.cells[2].text = str(4 + random.random())[0:4]
            j.cells[3].text = f".{random.randint(3, 9)}{random.randint(0, 9)}{random.randint(0, 9)}"
        except:
            pass

        row += 1
doc.save("Tables_completed_check.docx")


import openpyxl
def main():
    wb = openpyxl.load_workbook("DATA.xlsx")
    ws = wb.active
    col = 1
    rowGap = 46
    for colVar in range(1,ws.max_column +1):
        row = 2
        ws.insert_cols(col+1)
        for rowVar in range(48,94):
            cellData = ws.cell(row=rowVar,column=col).value
            if type(cellData) == float:
                cellData = round(cellData,2)
            ws.cell(row=rowVar,column=col).value = None
            ws.cell(row=row,column=col+1).value = cellData
            row+=1
        col+=2
    wb.save("DATA2.xlsx")
main()
'''
import openpyxl
import random
doc = openpyxl.Workbook()
sheet = doc.active
for i in range(1,500):
    sheet.cell(row=i,column=1).value = float(str(4 + random.random())[0:4])
    sheet.cell(row=i,column=2).value = float(f"0.{random.randint(3, 9)}{random.randint(0, 9)}{random.randint(0, 9)}")
doc.save("DOCAAAA.xlsx")
