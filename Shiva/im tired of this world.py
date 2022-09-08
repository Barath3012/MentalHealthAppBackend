import openpyxl
import random

def mamamoomoo():
    doc = openpyxl.load_workbook("PSA_VOLUME.xlsx")
    sheet = doc.active
    col1 = col2 = 1
    nextRow = 1
    for i in range(1,sheet.max_row+1,3):
        for j in range(3):
            for k in range(sheet.cell(row=i+j,column=3).value):
                print(sheet.cell(row=i, column=1).value)
                if type(sheet.cell(row=i, column=1).value) == int:
                    sheet.cell(row=nextRow, column=5).value = random.randint(
                        int(str(sheet.cell(row=i, column=1).value)[0:2]),
                        int(str(sheet.cell(row=i, column=1).value)[
                            2:])) + round(random.random(), 1)
                    sheet.cell(row=nextRow, column=6).value = random.randint(
                        eval(sheet.cell(row=i + j, column=2).value)[0],
                        eval(sheet.cell(row=i + j, column=2).value)[1])
                    nextRow += 1
                    continue

                sheet.cell(row=nextRow, column=5).value = random.randint(eval(sheet.cell(row=i, column=1).value)[0],
                                                                       eval(sheet.cell(row=i, column=1).value)[
                                                                           1]) + round(random.random(), 1)
                sheet.cell(row=nextRow, column=6).value = random.randint(eval(sheet.cell(row=i + j, column=2).value)[0],
                                                                       eval(sheet.cell(row=i + j, column=2).value)[1])
                nextRow+=1
    doc.save("NYANPASU.xlsx")

mamamoomoo()

