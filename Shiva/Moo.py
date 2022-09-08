import openpyxl


def haha():
    doc = openpyxl.open("data_split.xlsx")
    docSheet = doc.active
    dummyVar = 1
    for startingCol in range(1,9,3):

        for i in range(1,docSheet.max_row + 1):
            currentValue = str(docSheet.cell(row=i,column=dummyVar).value)
            try:

                valueToBePlaced = [currentValue.rsplit("/")[0],currentValue.rsplit("/")[1]]
                print("Success")
            except Exception as e:
                print(e)

            for k in range(2):
                docSheet.cell(row=i, column=dummyVar+k+1).value = valueToBePlaced[k]

        dummyVar+=3

    doc.save("NOOOO.xlsx")

haha()