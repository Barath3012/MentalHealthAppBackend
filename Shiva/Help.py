import openpyxl

file = openpyxl.load_workbook("usdata.xlsx")
sheet = file.active
count = 1
for col in range(1,sheet.max_column+1):
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row,column=col).value == None:
            print(count)
            sheet.cell(row=row, column=col).value = "NA"
            count+=1

file.save("usdata.xlsx")