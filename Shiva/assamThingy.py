import openpyxl

def thingy1():

    file = openpyxl.load_workbook("bhumika_data_1.xlsx")
    sheetFile = file["correction household data sheet"]
    docToBeWritten = openpyxl.Workbook()

    for row in range(2,sheetFile.max_row + 1):

        if type(sheetFile.cell(row=row,column=1).value) == type(None):
            if sheetFile.cell(row=row,column=2).value not in docToBeWritten.sheetnames:
                docToBeWritten.create_sheet(sheetFile.cell(row=row,column=2).value)
            try:
                docToBeWrittenSheet = docToBeWritten[sheetFile.cell(row=row,column=2).value]
            except:
                print(sheetFile.cell(row=row,column=2).value)
                continue
            row2 = docToBeWrittenSheet.max_row + 1
            for col in range(1,sheetFile.max_column + 1):
                docToBeWrittenSheet.cell(row=max(row2,2),column=col).value = sheetFile.cell(row=row,column=col).value
            row2+=1

    docToBeWritten.save("Test1.xlsx")

def main():
    thingy1()

if __name__ == "__main__":
    main()