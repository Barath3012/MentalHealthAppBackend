
'''
day,month,year = eval(input("Enter the date : "))

if month != 2 and month < 13:
    if month % 2 == 0 and month > 7:
        if day <= 31:
            print("Valid date")
        else:
            print("Invalid date")
    elif month % 2 == 0 and month > 8:
        if day <= 30:
            print("Valid date")
        else:
            print("Invalid date")
    elif month % 2 == 0 and month < 8:
        if day <= 31:
            print("Valid date")
        else:
            print("Invalid date")
    elif month % 2 != 0 and month < 7:
        if day <= 30:
            print("Valid date")
        else:
            print("Invalid date")


else:
    if not (year % 4 == 0):
        if day < 30:
            print("Valid date")
        else:
            print("Invalid date")
    else:
        if year % 100 == 0:
            if year % 400 == 0:
                if day < 30:
                    print("Valid date")
                else:
                    print("Invalid date")

            else:
                if day < 29:
                    print("Valid date")
                else:
                    print("Invalid date")
        else:
            if day < 29:
                print("Valid date")

a,b = 0,1
n = int(input("Enter a limit"))
for i in range(1,n+1):
    d = a+b
    a = b
    b = d
    print(d)

n = int(input("Enter the number of lines"))
s = 1
for i in range(1,n+1):
    for j in range(1,i+1):
        print(s)
        s+=1

n = int(input("Enter the number of lines"))
for i in range(1,n+1):
    print(" " * (n-i) + "* " * i)
for i in range(n-1,0,-1):
    print(" " * (n-i)  + "* " * i)

datee = input("Do you wanna go out date me?")

if datee == "yes":
    print("ok i dont wanna go out with you anymore")

else:

    print("I like feet")'''

import openpyxl

def aa(file):
    doc = openpyxl.load_workbook(file)
    sheet = doc.active
    newDoc = openpyxl.Workbook()
    newSheet = newDoc.active
    colCount = 1
    for i in range(1,sheet.max_row+1):

        data = []
        for j in range(1,sheet.max_column+1,2):

            data = data + [j//2+1] * int(sheet.cell(row = i,column=j).value)

        b = 1
        for k in data:
            print(b, colCount)
            newSheet.cell(row=b,column=colCount).value = k
            print(newSheet.cell(row=b, column=colCount).value)
            b+=1
        colCount += 1
        data = []
        for l in range(2, sheet.max_column+1, 2):
            #print(int(sheet.cell(row=i, column=l).value),end= " ")
            data = data + [l // 2 ] * int(sheet.cell(row=i, column=l).value)


        b = 1
        for k in data:

            newSheet.cell(row=b, column=colCount).value = k

            b += 1
        colCount += 1
    newDoc.save("HAHA.xlsx")

aa("book4.xlsx")


















