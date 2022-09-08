import openpyxl
import pywhatkit
doc = openpyxl.load_workbook(r"C:\Users\barat\OneDrive\Documents\yeay.xlsx")
sgeet=doc.active
a = ""
import os

j=10
for i in range(2,sgeet.max_row+1):
    a+= str(sgeet.cell(row=i,column=1).value).replace(" ","") + (", " if i%10!=0 else "")
    if i%10==0:
        print(a)
        a=""