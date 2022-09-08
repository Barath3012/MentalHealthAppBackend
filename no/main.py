
import csv
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

def group_add(name):
    l=[]
    with open('Samples/people.csv','rt')as f:
      data = csv.reader(f)
      for row in data:
          l.append(row[0])
    driver = webdriver.Chrome(executable_path="chromedriver.exe")
    driver.maximize_window()
    r=driver.get('https://web.whatsapp.com/')
    time.sleep(20)
    try:
        driver.find_element(by=By.XPATH,value='//*[@title="{}"]'.format(name)).click()# name of the group
    except:
        print("WhatsApp group doesn't exist")
        print("Please Try again")
        return
    f = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@class="_3V5x5"]')))
    f.click()
    f = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@class="_3p0T6"]')))
    f.click()
    print(f)
    '''
    for i in l :
        driver.find_element(by=By.XPATH,value='//*[@title="Search…"]').send_keys(i)
        time.sleep(5)
        driver.find_element(by=By.XPATH,value='//*[@title="{}"]'.format(i)).click()

    driver.find_element(by=By.XPATH,value='//*[@data-icon="checkmark-light"]').click()
    time.sleep(5)
    f = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH,'//*[@class="_2eK7W _3PQ7V"]')))
    f.click()  '''
    return
#group_add('test')

def sayhi(name):
    doc = openpyxl.load_workbook("Samples/file-1.xlsx")
    sheet=doc.active
    l = []
    m = []
    n = []
    for i in range(2,sheet.max_row+1):
        if sheet.cell(row=i,column=1).value!=None and sheet.cell(row=i,column=1).value!='':
            l.append(int(str(sheet.cell(row=i,column=1).value).lstrip("+").lstrip("9").lstrip("1").replace(" ","")))

    for i in range(2,sheet.max_row+1):
        if int(str(sheet.cell(row=i,column=2).value).replace(" ",""))not in l:
            m.append(int(str(sheet.cell(row=i,column=2).value).replace(" ","")))
    for i in range(2,sheet.max_row+1):

        n.append(int(str(sheet.cell(row=i,column=2).value).replace(" ","")))



    print(m)
    print(l)
    print(n)
    print(len(l))
    print(len(m))
sayhi(1)