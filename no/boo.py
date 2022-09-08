import openpyxl

def foo(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    wbook = openpyxl.Workbook()
    wsheet = wbook.active

    sheetmax = 3

    for i in range(8,sheet.max_row+1):
        print(i)
        #listvar = []
        datesThisTime = ["-","-","-"]
        maxThisRow = 0
        maxThisRowNo = 0
        dateRn = "-"
        for j in range(4,sheet.max_column):
            #print(j,"j")
            try:
                #listvar.append(float(sheet.cell(row=i,column=j).value))
                if maxThisRow < float(sheet.cell(row=i,column=j).value):
                    maxThisRow = float(sheet.cell(row=i,column=j).value)
                    maxThisRowNo = j
                    dateRn = str(sheet.cell(row=5,column=j).value)[:11]
            except:
                #print(sheet.cell(row=i,column=j).value,i,j)
                pass
        #maxThisRow = max(listvar)

        if maxThisRow != 0:
            datesThisTime[0] = dateRn

        predictMin = (60/100) * maxThisRow
        difference = 0
        leastDifference = predictMin
        newMin = 0
        newMinRowNo = 0
        dateRn = "-"
        for k in range(maxThisRowNo,sheet.max_column):
            #print(k,"k")
            try:
                difference = predictMin - float(sheet.cell(row=i,column=k).value)
                if difference < 0 :
                    continue
                #print(difference,leastDifference)
                if difference < leastDifference:
                    leastDifference = difference
                    newMin = float(sheet.cell(row=i,column=k).value)
                    newMinRowNo = k
                    dateRn = str(sheet.cell(row=5,column=k).value)[:11]
            except:
                pass
        if newMin != 0:
            datesThisTime[1] = dateRn

        #print(newMin,newMinRowNo)
        newMax = newMaxRowNo = 0
        difference = 0
        predictMax = newMin * (140/100)
        leastDifference = predictMax

        dateRn = "-"
        for l in range(newMinRowNo,sheet.max_column):
            try:
                difference = float(sheet.cell(row=i,column=l).value) - predictMax
                if difference < 0:
                    continue
                #print(difference,leastDifference)
                if difference < leastDifference:
                    leastDifference = difference
                    newMax = float(sheet.cell(row=i,column=l).value)
                    newMaxRowNo = l
                    dateRn = str(sheet.cell(row=5, column=l).value)[:11]
            except:
                pass
        if newMax != 0:
            datesThisTime[2] = dateRn

        wsheet.cell(row=i, column=sheetmax + 1).value = maxThisRow
        wsheet.cell(row=i, column=sheetmax + 2).value = datesThisTime[0]
        wsheet.cell(row=i, column=sheetmax + 3).value = newMin
        wsheet.cell(row=i, column=sheetmax + 4).value = datesThisTime[1]
        wsheet.cell(row=i, column=sheetmax + 5).value = newMax
        wsheet.cell(row=i, column=sheetmax + 6).value = datesThisTime[2]
        #sheet.cell(row=i,column=sheetmax+1).value = maxThisRow
        #sheet.cell(row=i, column=sheetmax+2).value = datesThisTime[0]
        #sheet.cell(row=i, column=sheetmax+3).value = newMin
        #sheet.cell(row=i, column=sheetmax+4).value = datesThisTime[1]
        #sheet.cell(row=i, column=sheetmax+5).value = newMax
        #sheet.cell(row=i, column=sheetmax+6).value = datesThisTime[2]

    for i in range(1,4):
        for j in range(1,sheet.max_row+1):
            wsheet.cell(row=j,column=i).value = sheet.cell(row=j,column=i).value

    wbook.save("hahaoutput2.xlsx")


if __name__ == "__main__":
    foo("haha2.xlsx")
